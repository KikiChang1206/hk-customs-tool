import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import math

# 1. 網頁基本設定
st.set_page_config(page_title="HK報關文件轉換器", layout="centered")

st.markdown("""
    <style>
    .stApp { background-color: #0E1117; }
    .big-title { font-size: 30px !important; font-weight: bold; color: #FFFFFF !important; }
    .stFileUploader section { background-color: #FFFFFF !important; border-radius: 10px; }
    div.stButton > button { background-color: #FFFFFF !important; color: #000000 !important; border: 2px solid #000000 !important; height: 50px; font-weight: bold; width: 100%; }
    .stMarkdown p, label { color: #FFFFFF !important; }
    </style>
    """, unsafe_allow_html=True)

st.markdown('<p class="big-title">🇭🇰 HK 報關文件轉換器</p>', unsafe_allow_html=True)

# 2. 檔案上傳
uploaded_files = st.file_uploader("請一次拖入所有 4 個必要檔案", type=['xls', 'xlsx'], accept_multiple_files=True)

files_dict = {"Invoice": None, "Packing": None, "北方文件": None, "OrderList": None}

if uploaded_files:
    for f in uploaded_files:
        fname = f.name.lower()
        if "invoice" in fname: files_dict["Invoice"] = f
        elif "packing" in fname: files_dict["Packing"] = f
        elif "manifest" in fname or "北方" in fname: files_dict["北方文件"] = f
        elif "order" in fname: files_dict["OrderList"] = f

# 3. 狀態顯示
st.write("### 📋 檔案讀取狀態")
c1, c2 = st.columns(2)
with c1:
    st.markdown(f"{'✅' if files_dict['Invoice'] else '❌'} **Invoice**")
    st.markdown(f"{'✅' if files_dict['Packing'] else '❌'} **Packing**")
with c2:
    st.markdown(f"{'✅' if files_dict['北方文件'] else '❌'} **北方文件**")
    st.markdown(f"{'✅' if files_dict['OrderList'] else '❌'} **Order List**")

# 4. 執行轉換
if all(files_dict.values()):
    if st.button("🚀 執行品牌校驗與轉換", use_container_width=True):
        try:
            def smart_read_excel(file_obj, **kwargs):
                if file_obj.name.endswith('.xls'): return pd.read_excel(file_obj, engine='xlrd', **kwargs)
                else: return pd.read_excel(file_obj, engine='openpyxl', **kwargs)

            # --- A. 預讀取品牌校驗 ---
            df_inv_raw = smart_read_excel(files_dict["Invoice"], header=None, dtype=str).fillna('')
            df_pac_raw = smart_read_excel(files_dict["Packing"], header=None, dtype=str).fillna('')
            df_n_export = smart_read_excel(files_dict["北方文件"], sheet_name='出口明細', dtype=str).fillna('')
            df_n_bag_raw = smart_read_excel(files_dict["北方文件"], sheet_name='袋數編號', dtype=str)
            df_order = smart_read_excel(files_dict["OrderList"], dtype=str).fillna('')

            inv_name = str(df_inv_raw.iloc[1, 0])
            pac_name = str(df_pac_raw.iloc[1, 0])
            north_remark = str(df_n_export.iloc[0, 8]) 

            brands = {
                "蜜凱": {"eng": "COSMETICS", "label": "TRUU+TRUE YOU", "key": "蜜凱", "suffix": "蜜凱"},
                "歐瑞": {"eng": "food supplement", "label": "ALLRE", "key": "歐瑞", "suffix": "歐瑞"},
                "綺麗絲": {"eng": "MAKEUP", "label": "MKUP", "key": "綺麗絲", "suffix": "綺麗絲"}
            }

            current_brand = None
            for b_name, b_info in brands.items():
                if b_info["key"] in inv_name:
                    current_brand = b_info
                    break

            if not current_brand:
                st.error("❌ 無法辨識公司品牌（蜜凱/歐瑞/綺麗絲），請檢查 Invoice A2 內容。")
                st.stop()

            # 交叉檢查
            key = current_brand["key"]
            if key not in pac_name or key not in north_remark:
                st.error(f"🚨 檔案品牌不匹配！偵測到主品牌為：【{key}】")
                st.warning(f"Packing 內容: {pac_name}")
                st.warning(f"北方文件備註: {north_remark}")
                st.stop()

            # --- B. 執行正式處理 ---
            with st.spinner(f'正在處理品牌：{current_brand["suffix"]} ...'):
                tw_now = datetime.utcnow() + timedelta(hours=8)
                t_str = tw_now.strftime("%Y%m%d")
                final_filename = f"{t_str}_HK_{current_brand['suffix']}.xlsx"

                wb = Workbook(); ws = wb.active; ws.title = "HK最終報關檔"

                # 欄寬與行高預設
                ws.column_dimensions['A'].width = 5.5
                col_widths = {'B': 20.8, 'C': 19.2, 'D': 14.7, 'E': 14, 'F': 14, 'G': 8.7, 'H': 13, 'I': 51.82, 'J': 30, 'K': 17.9, 'L': 8.7, 'M': 8.7, 'N': 8.09, 'O': 10.91, 'P': 10.5, 'Q': 8.09}
                for col, width in col_widths.items(): ws.column_dimensions[col].width = width
                
                # 表頭行高固定
                ws.row_dimensions[1].height = 77; ws.row_dimensions[2].height = 25.2
                for r in range(3, 7): ws.row_dimensions[r].height = 12.5
                ws.row_dimensions[7].height = 49.5; ws.row_dimensions[8].height = 25.2
                for r in range(9, 13): ws.row_dimensions[r].height = 12.5

                # 表頭填充邏輯 (同前版)
                def get_inv(ref):
                    cmap = {'A':0, 'B':1, 'C':2, 'D':3, 'E':4, 'F':5, 'G':6, 'H':7, 'I':8}
                    c = cmap[ref[0]]; r = int(ref[1:]) - 1
                    return df_inv_raw.iloc[r, c]

                ws["B1"] = "INVOICE/PACKING"; ws["B1"].font = Font(name='Arial', size=28, bold=True); ws.merge_cells("B1:E1")
                ws["B1"].alignment = Alignment(horizontal='left', vertical='center')
                
                head_configs = [("B2", get_inv("A2"), "B2:I2", 10, False, True), ("B3", get_inv("A3"), "B3:E3", 10, False, False), ("F3", get_inv("E3"), "F3:I3", 10, False, False), ("B4", get_inv("A4"), "B4:I4", 10, False, False), ("B5", get_inv("A5"), "B5:E5", 10, False, False), ("F5", get_inv("E5"), "F5:I5", 10, False, False), ("B6", get_inv("A6"), "B6:I6", 10, False, False), ("B7", get_inv("A7"), "B7:E7", 10, False, False), ("F7", get_inv("E7"), "F7:I7", 10, False, True), ("B8", get_inv("A8"), "B8:E8", 10, False, True), ("F8", get_inv("E8"), "F8:I8", 10, False, False), ("B9", get_inv("A9"), "B9:D9", 10, False, False), ("E9", get_inv("D9"), "E9:G9", 10, False, False), ("H9", get_inv("G9"), "H9:I9", 10, False, False), ("B10", get_inv("A10"), "B10:E10", 10, False, False), ("F10", get_inv("E10"), "F10:I10", 10, False, False)]
                for c_id, cont, m_range, sz, bld, wrp in head_configs:
                    ws[c_id] = cont; ws[c_id].font = Font(name='Arial', size=sz, bold=bld)
                    ws[c_id].alignment = Alignment(wrap_text=wrp, vertical='center'); ws.merge_cells(m_range)

                ws['B11'] = "FOB"; ws['B11'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws['B11'].font = Font(name='Arial', size=10, bold=True)

                # 標題列
                headers = ["提單編號", "訂單編號", "好馬吉袋號", "條碼", "單箱重量(GW)", "品項淨重", "品項英文名稱", "品項中文名稱", "品項備註", "品項品牌", "品項產地", "品項數量", "單位", "品項單價", "品項小計", "幣別"]
                green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for i, title in enumerate(headers, 2):
                    cell = ws.cell(row=13, column=i, value=title); cell.fill = green_fill; cell.font = Font(name='Arial', size=10, bold=True)
                    cell.border = thin_border; cell.alignment = Alignment(horizontal='center', vertical='center')

                # 字典建置
                barcode_dict = df_n_bag_raw.fillna('').set_index(df_n_bag_raw.columns[0])[df_n_bag_raw.columns[1]].to_dict()
                bag_dict = df_n_export.set_index(df_n_export.columns[1])[df_n_export.columns[6]].to_dict()
                bag_count = len(df_n_bag_raw[df_n_bag_raw.iloc[:, 1].str.strip() != ""])

                # 資料處理
                all_rows = []
                for _, r in df_order.iterrows():
                    hawb, oid = str(r.iloc[1]).strip(), str(r.iloc[3]).strip()
                    bag_no = bag_dict.get(hawb, ""); barcode = barcode_dict.get(bag_no, "")
                    gw_raw = r.iloc[30]
                    try: gw_num = float(gw_raw)
                    except: gw_num = 0.0
                    all_rows.append({"hawb": hawb, "oid": oid, "bag_no": bag_no, "barcode": barcode, "gw_raw": gw_raw, "gw_num": gw_num, "orig_row": r})

                all_rows.sort(key=lambda x: (x["barcode"], x["hawb"], x["gw_num"]))

                prev_hawb, curr_row, item_no = None, 14, 1
                sum_gw = sum_nw = sum_qty = sum_amount = 0.0

                for entry in all_rows:
                    ws.cell(row=curr_row, column=1, value=item_no).font = Font(name='Arial', size=10)
                    ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    
                    hawb, oid, bag_no, barcode, gw_raw = entry["hawb"], entry["oid"], entry["bag_no"], entry["barcode"], entry["gw_raw"]
                    r = entry["orig_row"]

                    gw_display = nw_display = ""
                    if hawb != prev_hawb:
                        gw_display = gw_raw
                        try:
                            val_gw = float(gw_raw); sum_gw += val_gw
                            calc_nw = val_gw - 0.2
                            final_nw = calc_nw if calc_nw >= 0.01 else 0.01
                            nw_display = "{:.2f}".format(final_nw); sum_nw += final_nw
                        except: pass
                    
                    try: sum_qty += float(r.iloc[37]); sum_amount += float(r.iloc[40])
                    except: pass

                    row_content = [hawb, oid, bag_no, barcode, gw_display, nw_display, current_brand["eng"], r.iloc[33], r.iloc[34], current_brand["label"], r.iloc[36], r.iloc[37], "SET", r.iloc[39], r.iloc[40], "TWD"]
                    
                    # --- 計算此行需要的行高 ---
                    max_lines = 1
                    for col_idx, val in enumerate(row_content, 2):
                        c = ws.cell(row=curr_row, column=col_idx, value=val)
                        c.font = Font(name='Arial', size=10)
                        c.border = thin_border
                        
                        # 設定換行與對齊
                        if col_idx in [9, 10]: # 品項中文名稱 與 品項備註 欄位
                            c.alignment = Alignment(wrap_text=True, vertical='center')
                            # 動態估計行數: 內容長度 / (欄寬 * 0.8) -> 0.8 是粗略估計中文字在 Arial 10號下的佔比
                            current_col_width = col_widths.get(chr(64 + col_idx), 10)
                            text_str = str(val) if val else ""
                            # 考慮中文字元與手動換行符
                            lines = text_str.count('\n') + 1
                            chars_per_line = max(1, math.floor(current_col_width * 0.75)) # 經驗係數
                            wrap_lines = math.ceil(len(text_str) / chars_per_line)
                            max_lines = max(max_lines, lines, wrap_lines)
                        else:
                            c.alignment = Alignment(vertical='center')

                    # 設定行高 (1行大約 15~18 點高)
                    ws.row_dimensions[curr_row].height = max_lines * 16.5
                    
                    prev_hawb, curr_row, item_no = hawb, curr_row + 1, item_no + 1

                # 統計匯總 (同前版)
                packing_last_val = df_pac_raw.iloc[-1, 0] if not df_pac_raw.empty else ""
                ws.cell(row=curr_row, column=2, value=packing_last_val).font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=5, value=f"包{bag_count}袋").font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=6, value="{:.2f}".format(sum_gw)).font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=7, value="{:.2f}".format(sum_nw)).font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=13, value=sum_qty).font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=14, value="SET").font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=16, value="{:.2f}".format(sum_amount)).font = Font(name='Arial', size=10, bold=True)
                ws.cell(row=curr_row, column=17, value="TWD").font = Font(name='Arial', size=10, bold=True)

                output = BytesIO(); wb.save(output); st.balloons()
                st.success(f"✅ 品牌【{current_brand['suffix']}】處理完成！")
                st.download_button(label=f"📥 下載 {final_filename}", data=output.getvalue(), file_name=final_filename, use_container_width=True)

        except Exception as e:
            st.error(f"發生錯誤：{e}")
